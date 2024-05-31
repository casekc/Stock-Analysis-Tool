import os
import shutil
import tempfile
import zipfile
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.xml.constants import ARC_SHARED_STRINGS, SHEET_MAIN_NS
from openpyxl.xml.functions import Element, SubElement, tostring
from lxml.etree import ElementTree
from openpyxl import Workbook
from openpyxl.xml.constants import SHARED_STRINGS
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from lxml.etree import Element, SubElement, tostring
from openpyxl.xml.functions import tostring
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.xml.functions import Element, SubElement
from openpyxl.xml.functions import tostring
from openpyxl.xml.functions import tostring
from lxml.etree import Element, SubElement, tostring



import pandas as pd

import xlrd

from extractor import stockInput

def analyzeNetIncome(ticker_list):
    temp_dir = tempfile.mkdtemp()

    try:
        temp_filepath = os.path.join(temp_dir, 'temp_file.xlsx')

        shutil.copy(input_filepath, temp_filepath)

        wb = load_workbook(filename=temp_filepath)

        ws = wb.active
        ws['F1'] = 'G1'
        ws['G1'] = 'G2'
        ws['H1'] = 'G3'
        ws['I1'] = 'AAGR'
        for i, ticker in enumerate(ticker_list, start=2):
            row_num = str(i)
            ws['F' + row_num] = '=((D' + row_num + '- E' + row_num + ')/E' + row_num + ')*100'
            ws['G' + row_num] = '=((C' + row_num + '- D' + row_num + ')/D' + row_num + ')*100'
            ws['H' + row_num] = '=((B' + row_num + '- C' + row_num + ')/C' + row_num + ')*100'
            ws['I' + row_num] = '=AVERAGE(F' + row_num + ':H' + row_num + ')'

        wb.save(output_filepath)

        create_shared_strings_xml(output_filepath, wb.shared_strings)

        print("Analysis completed successfully.")

        with zipfile.ZipFile(output_filepath, 'r') as archive:
            file_list = archive.namelist()
            if 'xl/sharedStrings.xml' in file_list:
                print("xl/sharedStrings.xml is in the output Excel archive.")
            else:
                print("xl/sharedStrings.xml is NOT in the output Excel archive.")

    except Exception as e:
        print(f"Error analyzing net income: {str(e)}")

    finally:
        shutil.rmtree(temp_dir)





def generate_shared_strings_xml(shared_strings):
    xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>' \
          '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{0}" uniqueCount="{0}">'.format(
        len(shared_strings))
    for shared_string in shared_strings:
        xml += '<si><t>{}</t></si>'.format(shared_string)
    xml += '</sst>'
    return xml








def analyzeHistoricalPrices(ticker_list):
    for ticker in ticker_list:
        wb = load_workbook(xlsx_path + ticker + '.xlsx')
        ws = wb.active

        length = ws.max_row

        ws['G1'] = 'Pco'
        ws['H1'] = 'Phl'
        ws['I1'] = 'AvgPco'
        ws['J1'] = 'AvgPhl'

        for i in range(2, length + 1):
            i = str(i)
            ws['G' + i] = '=((B' + i + '-D' + i + ')/(D' + i + '))*100'  # B2-D2
            ws['H' + i] = '=((E' + i + '-F' + i + ')/(F' + i + '))*100'  # E2-F2

        ws['I2'] = '=AVERAGE(G2:G' + str(length) + ')'
        ws['J2'] = '=AVERAGE(H2:H' + str(length) + ')'

        wb.save(output_directory + ticker + '.xlsx')





