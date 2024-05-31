import os
import pandas as pd
import numpy as np
import xlrd
import openpyxl
from openpyxl import load_workbook, Workbook
from pyxlsb import open_workbook as open_xlsb
import xlwings as xw
import pythoncom

def comparativeAnalysis():
    wb = openpyxl.load_workbook(net_income_analysis_path, data_only=True)
    sheet = wb.active

    aagr_values = [cell.value for cell in sheet['I'][1:]]

    tickers = [cell.value for cell in sheet['A'][1:]]

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    new_sheet['A1'] = 'Ticker'
    new_sheet['B1'] = 'AAGR'
    new_sheet['C1'] = 'PPI'
    new_sheet['D1'] = 'VI'

    for i, (ticker, aagr) in enumerate(zip(tickers, aagr_values), start=2):
        new_sheet['A{}'.format(i)] = ticker
        new_sheet['B{}'.format(i)] = aagr




    new_wb.save(comparative_analysis_path)
    data = []
    columns = ['PPI', 'VI']

    for i in os.listdir(historical_price_analysis_path):
        wb = openpyxl.load_workbook(historical_price_analysis_path + i, data_only=True)
        ws = wb.active
        AvgPco = ws['I2']
        AvgPhl = ws['J2']
        data.append([AvgPco.value, AvgPhl.value])

    df = pd.DataFrame(data, columns=columns)

    print(df)

    existing_wb = openpyxl.load_workbook()
    #writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = existing_wb
    writer.sheets = dict((ws.title, ws) for ws in existing_wb.worksheets)
    df.to_excel(writer, index=False, sheet_name='Sheet', startcol=2, startrow=1, header=False)
    writer.save()
