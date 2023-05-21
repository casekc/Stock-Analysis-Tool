import os
import pandas as pd
import numpy as np
import xlrd
import openpyxl
from openpyxl import load_workbook, Workbook
from pyxlsb import open_workbook as open_xlsb
import xlwings as xw
import pythoncom

new_path = r'C:\Users\cummi\Desktop\webscrap\bin\converted_files\net_income_analysis.xlsx'

def old_evaluate_formula(cell):
    # Evaluate the formula in the cell and return the calculated value
    calculated_value = cell.parent.parent.calculate_cell(cell.coordinate).value
    return calculated_value

def comparativeAnalysis():
    net_income_analysis_path = r'C:\Users\cummi\Desktop\webscrap\bin\converted_files\net_income_analysis.xlsx'
    comparative_analysis_path = r'C:\Users\cummi\Desktop\webscrap\bin\comparative_analyses\comparative_analysis.xlsx'

    # Load the net income analysis workbook using openpyxl with data_only option
    wb = openpyxl.load_workbook(net_income_analysis_path, data_only=True)
    sheet = wb.active

    # Get the calculated values of AAGR column
    aagr_values = [cell.value for cell in sheet['I'][1:]]

    # Extract the tickers and AAGR values
    tickers = [cell.value for cell in sheet['A'][1:]]

    # Create a new workbook and sheet using openpyxl
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    # Write the extracted data to the new sheet
    new_sheet['A1'] = 'Ticker'
    new_sheet['B1'] = 'AAGR'
    for i, (ticker, aagr) in enumerate(zip(tickers, aagr_values), start=2):
        new_sheet['A{}'.format(i)] = ticker
        new_sheet['B{}'.format(i)] = aagr

    # Save the new workbook
    new_wb.save(comparative_analysis_path)
old_old_old_comparativeAnalysis()