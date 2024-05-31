import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.xml.constants import SHARED_STRINGS
from extractor import stockInput
import pathlib
import csv
import os

def fix_excel_files(hpa_directory, nia_directory):
    for filename in os.listdir(hpa_directory):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(hpa_directory, filename)
            try:
                wb = load_workbook(filepath, read_only=True)
                fixed_wb = load_workbook(filepath)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    fixed_sheet = fixed_wb[sheet_name]
                    fixed_sheet.delete_rows(1, fixed_sheet.max_row)
                    for row in sheet.iter_rows(values_only=True):
                        fixed_sheet.append(row)
                fixed_wb.save(filepath)
                print(f"Fixed file: {filename}")
            except Exception as e:
                print(f"Error fixing file: {filename}\n{str(e)}")
    for filename in os.listdir(nia_directory):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(nia_directory, filename)
            try:
                wb = load_workbook(filepath, read_only=True)
                fixed_wb = load_workbook(filepath)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    fixed_sheet = fixed_wb[sheet_name]
                    fixed_sheet.delete_rows(1, fixed_sheet.max_row)
                    for row in sheet.iter_rows(values_only=True):
                        fixed_sheet.append(row)
                fixed_wb.save(filepath)
                print(f"Fixed file: {filename}")
            except Exception as e:
                print(f"Error fixing file: {filename}\n{str(e)}")






def convertCSV():
    for filename in os.listdir(directory):
        if filename.endswith('.csv'):
            file_path = os.path.join(directory, filename)

            df = pd.read_csv(file_path)

            export_filename = os.path.splitext(filename)[0] + '.xlsx'
            export_path = os.path.join(export_directory, export_filename)
            df.to_excel(export_path, index=False)


def renameCSV(ticker_list):
    csv_path = pathlib.Path()
    generator = csv_path.iterdir()
    dir_list = list(generator)
    print(dir_list)

    for item in dir_list:
        for ticker in ticker_list:
            path = str(item)  
            new_path = path[:40] + ticker + '.csv'

            if item.is_file():
                new_file_path = csv_path / new_path
                if not new_file_path.exists(): 
                    item.rename(new_file_path)
                    print(f"Renamed {item} to {new_path}")
                else:
                    print(f"Skipping {item}. Destination file {new_path} already exists.")
            else:
                print(f"Skipping {item} as it is not a file")


def format_list(financials_list):
    numbers_list = []
    for i in financials_list:
        numbers = i[10:].split('$')[1:]
        numbers_list.append(numbers)
    return numbers_list


def move_hyphen(nested_list):
    result = []
    for sublist in nested_list:
        new_sublist = []
        for element in sublist:
            if element.endswith("-"):
                element = element[:-1] 
                if new_sublist and new_sublist[-1].endswith("-"):
                    new_sublist[-1] += element
                else:
                    new_sublist.append(
                        element)
                new_sublist.append("-")
            else:
                new_sublist.append(element)
        result.append(new_sublist)
    return result


def split_list(data):
    sublists = []
    sublist = []
    total_revenue_count = 0

    for item in data:
        if item == "Total Revenue":
            total_revenue_count += 1
            if total_revenue_count > 2:
                sublists.append(sublist)
                sublist = []
        sublist.append(item)

    sublists.append(sublist)

    return sublists


def extract_net_income(data):
    extracted_data = []
    for sublist in data:
        if 'Net Income' in sublist:
            index = sublist.index('Net Income')
            extracted_data.append(sublist[index:index + 5])
    return extracted_data
